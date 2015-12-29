# -*- coding: utf-8 -*-
"""
Created on Sun Jul 12 16:19:27 2015

@author: Phenomics
"""

from openpyxl import load_workbook

#%% input the data
data_file_name = 'Homework-BusinessAnalysisML.xlsx';
workbook = load_workbook(data_file_name, use_iterators=True)

#%% read data label
labels = dict();
sheet3 = workbook.get_sheet_names()[2];
worksheet3 = workbook.get_sheet_by_name(sheet3);
for row in list(worksheet3.iter_rows())[1:]:
    col1 = row[0].internal_value.split('-');
    key = col1[0].lower();        
    if row[1].internal_value=="Bad":
        labels[key]=labels.get(key,0)-1
    else:
        labels[key]=labels.get(key,0)+1

#%%
OutputCSVFile = "datasheet.csv"

def appendScoreColumn(rowData,isheader=False):
    res=[]
    for cell in rowData:
        res.append(cell.internal_value)
    if isheader:
        res.append("loan performance")
    else:
        res.append(labels.get(rowData[0].internal_value,'unknown'))
    return res

sheet2 = workbook.get_sheet_names()[1];
worksheet2 = workbook.get_sheet_by_name(sheet2);
listWs2 = list(worksheet2.iter_rows())
outList=[]
header = listWs2[0]
outList.append(appendScoreColumn(header,True))
for row in listWs2[1:]:
    outList.append(appendScoreColumn(row,False))

with open(OutputCSVFile,'w') as f:
    for row in outList:
        f.write(','.join(map(str,row))+'\n')

    

